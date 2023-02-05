# /**********************************************************************
# *                                                                     *
# * Copyright (c) 2021 Javier Braña <javier.branagutierrez@gmail.com>  *
# *                                                                     *
# * This program is free software; you can redistribute it and/or modify*
# * it under the terms of the GNU Lesser General Public License (LGPL)  *
# * as published by the Free Software Foundation; either version 2 of   *
# * the License, or (at your option) any later version.                 *
# * for detail see the LICENCE text file.                               *
# *                                                                     *
# * This program is distributed in the hope that it will be useful,     *
# * but WITHOUT ANY WARRANTY; without even the implied warranty of      *
# * MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the       *
# * GNU Library General Public License for more details.                *
# *                                                                     *
# * You should have received a copy of the GNU Library General Public   *
# * License along with this program; if not, write to the Free Software *
# * Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307*
# * USA                                                                 *
# *                                                                     *
# ***********************************************************************


import FreeCAD
import Part
import os

if FreeCAD.GuiUp:
    import FreeCADGui, os
    from PySide import QtCore, QtGui
    from PySide.QtCore import QT_TRANSLATE_NOOP
else:
    # \cond
    def translate(ctxt, txt):
        return txt
    def QT_TRANSLATE_NOOP(ctxt, txt):
        return txt
    # \endcond

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, GradientFill, Font
import PVPlantResources
import PVPlantSite

# Estilos:
thin = Side(border_style="thin", color="7DA4B8")
double = Side(border_style="double", color="ff0000")
border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)
border_fat = Border(top=thin, left=thin, right=thin, bottom=thin)
# fill = PatternFill("solid", fgColor="DDDDDD")
# fill = GradientFill(stop=("000000", "FFFFFF"))

scale = 0.001

def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill

def spreadsheetBOQFrames(sheet, sel):
    sheet['A1'] = 'Index'
    sheet['B1'] = 'Frame'
    sheet['C1'] = 'X'
    sheet['D1'] = 'Y'
    sheet['E1'] = 'Z'
    sheet['F1'] = 'Angle N-S'
    sheet['G1'] = 'Angle L-W'
    sheet['H1'] = 'Nº Poles'

    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 15
    sheet.row_dimensions[1].height = 40

    style_range(sheet, 'A1:H1',
                border=Border(top=thin, left=thin, right=thin, bottom=thin),
                fill=PatternFill("solid", fgColor="7DA4B8"),
                font=Font(name='Gill Sans MT', size=10, b=True, color="FFFFFF"),
                alignment=Alignment(horizontal="center", vertical="center"))

    for ind in range(0, len(sel)):
        row = ind + 2
        sheet['A{0}'.format(row)] = ind + 1
        sheet['B{0}'.format(row)] = sel[ind].Label
        sheet['C{0}'.format(row)] = sel[ind].Placement.Base.x * scale
        sheet['D{0}'.format(row)] = sel[ind].Placement.Base.y * scale
        sheet['E{0}'.format(row)] = sel[ind].Placement.Base.z * scale
        sheet['F{0}'.format(row)] = sel[ind].Placement.Rotation.toEuler()[0]
        sheet['G{0}'.format(row)] = sel[ind].Placement.Rotation.toEuler()[1]
        sheet['H{0}'.format(row)] = sel[ind].NumberPole.Value
        style_range(sheet, 'A' + str(row) + ':H' + str(row),
                    border=Border(top=thin, left=thin, right=thin, bottom=thin),
                    font=Font(name='Gill Sans MT', size=10),
                    alignment=Alignment(horizontal="center", vertical="center"))

def spreadsheetBOQPoles(sheet, sel):
    import MeshPart as mp
    # Headers:
    sheet['A1'] = 'Frame'
    sheet['B1'] = 'Pole'
    sheet['C1'] = 'X'
    sheet['D1'] = 'Y'
    sheet['E1'] = 'Z frame attach'
    sheet['F1'] = 'Z aerial head'
    sheet['G1'] = 'Pole length'
    sheet['H1'] = 'Pole aerial length'
    sheet['I1'] = 'Pole terrain enter length'

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 8
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.row_dimensions[1].height = 40
    style_range(sheet, 'A1:I1',
                border=Border(top=thin, left=thin, right=thin, bottom=thin),
                fill=PatternFill("solid", fgColor="7DA4B8"),
                font=Font(name='Gill Sans MT', size=11, b=True, color="FFFFFF"),
                alignment=Alignment(horizontal="center", vertical="center"))
    sheet['A2'] = ""
    sheet.row_dimensions[2].height = 5

    # Data:
    terrain = PVPlantSite.get().Terrain.Mesh #Shape
    terrain = FreeCAD.ActiveDocument.Mesh002.Mesh
    row = 3
    for frame_ind, frame in enumerate(sel):
        poles = frame.Shape.SubShapes[1].SubShapes
        group_from = row
        #frame_line = Part.LineSegment(poles[0].BoundBox.Center, poles[-1].BoundBox.Center)
        #frame_line_projection = terrain.makeParallelProjection(frame_line.toShape(), FreeCAD.Vector(0,0,1))

        for pole_ind, pole in enumerate(poles):
            zattach = .0
            center = pole.BoundBox.Center

            '''down = FreeCAD.Vector(center.x, center.y, terrain.BoundBox.ZMin)
            top = FreeCAD.Vector(center.x, center.y, terrain.BoundBox.ZMax)
            pole_axis = Part.LineSegment(top, down)
            for ed in frame_line_projection.Edges:
                tmp = Part.LineSegment(ed.Vertexes[0].Point, ed.Vertexes[1].Point)
                result = pole_axis.intersect(tmp)
                if len(result) > 0:
                    zattach = result[0].Z
                    break'''

            pp = mp.projectPointsOnMesh([center], terrain, FreeCAD.Vector(0, 0, 1))
            if len(pp) == 0:
                zattach = -999
            else:
                zattach = pp[0].z

            sheet['B{0}'.format(row)] = pole_ind + 1
            sheet['C{0}'.format(row)] = pole.Placement.Base.x * scale
            sheet['D{0}'.format(row)] = pole.Placement.Base.y * scale
            sheet['E{0}'.format(row)] = zattach * scale
            sheet['F{0}'.format(row)] = pole.BoundBox.ZMax * scale
            sheet['G{0}'.format(row)] = frame.PoleLength.Value * scale
            sheet['H{0}'.format(row)] = '=F{0}-E{0}'.format(row)
            sheet['I{0}'.format(row)] = '=G{0}-H{0}'.format(row)

            style_range(sheet, 'A' + str(row) + ':I' + str(row),
                        border=Border(top=thin, left=thin, right=thin, bottom=thin),
                        font=Font(name='Gill Sans MT', size=11,),
                        alignment=Alignment(horizontal="center", vertical="center"))
            row += 1

        style_range(sheet, 'A' + str(group_from) + ':F' + str(row - 1),
                    border=Border(top=thin, left=thin, right=thin, bottom=thin),
                    font=Font(name='Gill Sans MT', size=11,),
                    alignment=Alignment(horizontal="center", vertical="center"))
        sheet.merge_cells('A' + str(group_from) + ':A' + str(row - 1))
        cell = sheet['A' + str(group_from)]
        cell.value = sel[frame_ind].Label
        style_range(sheet, 'A' + str(group_from) + ':A' + str(row - 1),
                    border=Border(top=thin, left=thin, right=thin, bottom=thin),
                    font=Font(name='Gill Sans MT', size=11,),
                    alignment=Alignment(horizontal="center", vertical="center"))
        sheet['A{0}'.format(row)] = ""
        sheet.row_dimensions[row].height = 5
        row += 1

def spreadsheetBOQPanelCollision(sheet, sel):
    # Headers:
    sheet['A1'] = 'Frame'
    sheet['B1'] = 'Nombre'
    sheet['C1'] = 'X'
    sheet['D1'] = 'Y'
    sheet['E1'] = 'Z'
    sheet['G1'] = 'Ángulo E-O'
    sheet['H1'] = 'Nº Hincas'

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 8
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.row_dimensions[1].height = 40
    style_range(sheet, 'A1:I1',
                border=Border(top=thin, left=thin, right=thin, bottom=thin),
                fill=PatternFill("solid", fgColor="7DA4B8"),
                font=Font(name='Gill Sans MT', size=11, b=True, color="FFFFFF"),
                alignment=Alignment(horizontal="center", vertical="center"))
    sheet['A2'] = ""
    sheet.row_dimensions[2].height = 5

    # Data:
    for frame_ind in range(0, len(sel)):
        frame = sel[frame_ind]

        sheet['A{0}'.format(ind + 2)] = ind
        sheet['B{0}'.format(ind + 2)] = sel[ind].Label
        sheet['C{0}'.format(ind + 2)] = sel[ind].Placement.Base.x * scale
        sheet['D{0}'.format(ind + 2)] = sel[ind].Placement.Base.y * scale
        sheet['E{0}'.format(ind + 2)] = sel[ind].Placement.Base.z * scale
        sheet['F{0}'.format(ind + 2)] = sel[ind].Placement.Rotation.toEuler()[0]
        sheet['G{0}'.format(ind + 2)] = sel[ind].Placement.Rotation.toEuler()[1]
        sheet['H{0}'.format(ind + 2)] = sel[ind].NumberPole.Value
    mywb.save(filename)


class _CommandBOQMechanical:
    def GetResources(self):
        return {'Pixmap': str(os.path.join(PVPlantResources.DirIcons, "boqm.svg")),
                'Accel': "R, M",
                'MenuText': QT_TRANSLATE_NOOP("Placement", "BOQ Mecánico"),
                'ToolTip': QT_TRANSLATE_NOOP("Placement", "Calcular el BOQ de la")}

    def Activated(self):
        # make file global:
        sel = FreeCAD.ActiveDocument.findObjects(Name="Tracker")
        if len(sel) > 0:
            path = os.path.dirname(FreeCAD.ActiveDocument.FileName)
            filename = os.path.join(path, "BOQMechanical.xlsx")
            mywb = openpyxl.Workbook()
            sheet = mywb.active
            sheet.title = 'Frames information'
            spreadsheetBOQFrames(sheet, sel)

            sheet = mywb.create_sheet("Poles information")
            spreadsheetBOQPoles(sheet, sel)
            mywb.save(filename)

    def IsActive(self):
        if FreeCAD.ActiveDocument:
            return True
        else:
            return False

if FreeCAD.GuiUp:
    FreeCADGui.addCommand('BOQMechanical', _CommandBOQMechanical())
