import FreeCAD

if FreeCAD.GuiUp:
    import FreeCADGui, os
    from PySide import QtCore, QtGui
    from DraftTools import translate
    from PySide.QtCore import QT_TRANSLATE_NOOP
else:
    # \cond
    def translate(ctxt, txt):
        return txt

    def QT_TRANSLATE_NOOP(ctxt, txt):
        return txt
    # \endcond

try:
    _fromUtf8 = QtCore.QString.fromUtf8
except AttributeError:
    def _fromUtf8(s):
        return s


try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side, PatternFill, GradientFill, Font
    useexcel = True
except:
    useexcel = False

import PVPlantResources


def spreadsheetBOQFrames(sel):

    import os
    path = os.path.dirname(FreeCAD.ActiveDocument.FileName)
    filename = os.path.join(path, "example_filetest.xlsx")

    i0 = '{0}'
    f3 = '{:.3f}'
    scale = 0.001
    
    mywb = openpyxl.Workbook()
    #mywb = openpyxl.load_workbook('filetest.xlsx')
    sheet = mywb.active
    sheet.title = 'BOQ Frames'

    sheet['A1'] = 'Índice'
    sheet['B1'] = 'Nombre'
    sheet['C1'] = 'X'
    sheet['D1'] = 'Y'
    sheet['E1'] = 'Z'
    sheet['F1'] = 'Ángulo N-S'
    sheet['G1'] = 'Ángulo E-O'
    sheet['H1'] = 'Nº Hincas'

    #ws = wb.active
    #ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
    
    for ind in range(0, len(sel)):
        sheet['A{0}'.format(ind + 2)] = ind
        sheet['B{0}'.format(ind + 2)] = sel[ind].Label
        sheet['C{0}'.format(ind + 2)] = sel[ind].Placement.Base.x * scale
        sheet['D{0}'.format(ind + 2)] = sel[ind].Placement.Base.y * scale
        sheet['E{0}'.format(ind + 2)] = sel[ind].Placement.Base.z * scale
        sheet['F{0}'.format(ind + 2)] = sel[ind].Placement.Rotation.toEuler()[0]
        sheet['G{0}'.format(ind + 2)] = sel[ind].Placement.Rotation.toEuler()[1]
        sheet['H{0}'.format(ind + 2)] = sel[ind].NumberPole.Value

    mywb.save(filename)

    return
    import Spreadsheet



    sheet = FreeCAD.ActiveDocument.addObject("Spreadsheet::Sheet","BOQFrames")
    sheet.Label = "BOQ Frames"
    sheet.set('A1', 'Índice')
    sheet.set('B1', 'Nombre')
    sheet.set('C1', 'X')
    sheet.set('D1', 'Y')
    sheet.set('E1', 'Z')
    sheet.set('F1', 'Ángulo N-S')
    sheet.set('G1', 'Ángulo E-O')
    sheet.set('H1', 'Nº Hincas')

    for ind in range(0, len(sel)):
        sheet.set('A{0}'.format(ind + 2), str(ind))
        sheet.set('B{0}'.format(ind + 2), sel[ind].Label)
        sheet.set('C{0}'.format(ind + 2), f3.format(sel[ind].Placement.Base.x * scale))
        sheet.set('D{0}'.format(ind + 2), f3.format(sel[ind].Placement.Base.y * scale))
        sheet.set('E{0}'.format(ind + 2), f3.format(sel[ind].Placement.Base.z * scale))
        sheet.set('F{0}'.format(ind + 2), f3.format(sel[ind].Placement.Rotation.toEuler()[0]))
        sheet.set('G{0}'.format(ind + 2), f3.format(sel[ind].Placement.Rotation.toEuler()[1]))
        sheet.set('H{0}'.format(ind + 2), i0.format(sel[ind].NumberPole))

    sheet.recompute()

def spreadsheetBOQPoles(sel):
    import os
    path = os.path.dirname(FreeCAD.ActiveDocument.FileName)
    filename = os.path.join(path, "BOQPoles.xlsx")

    i0 = '{0}'
    f3 = '{:.3f}'
    scale = 0.001

    mywb = openpyxl.Workbook()
    # mywb = openpyxl.load_workbook('filetest.xlsx')
    sheet = mywb.active
    sheet.title = 'BOQ Poles'

    # Estilos:
    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    border_thin = Border(top = thin, left = thin, right = thin, bottom = thin)
    border_fat = Border(top = thin, left = thin, right = thin, bottom = thin)
    #fill = PatternFill("solid", fgColor="DDDDDD")
    #fill = GradientFill(stop=("000000", "FFFFFF"))
    #sheet['A1:B10']

    # Headers:
    sheet['A1'] = 'Índice'
    sheet.column_dimensions['A'].width = 8
    sheet['B1'] = 'Tracker'
    sheet.column_dimensions['B'].width = 30
    sheet['C1'] = 'Hinca'
    sheet.column_dimensions['C'].width = 8
    sheet['D1'] = 'X'
    sheet.column_dimensions['D'].width = 15
    sheet['E1'] = 'Y'
    sheet.column_dimensions['E'].width = 15
    sheet['F1'] = 'Z'
    sheet.column_dimensions['F'].width = 15
    sheet.row_dimensions[1].height = 40
    style_range(sheet, 'A1:F1',
                border = Border(top = thin, left = thin, right = thin, bottom = thin),
                fill = PatternFill("solid", fgColor="DDDDDD"),
                font = Font(b=True, color="FFFFFF"),
                alignment = Alignment(horizontal="center", vertical="center"))
    sheet['A2'] = ""
    sheet.row_dimensions[2].height = 5

    # Celdas de datos:
    row = 3
    for tracker_ind in range(1, len(sel)):
        poles = sel[tracker_ind].Shape.SubShapes
        group_from = row
        for pole_ind in range(0, int(sel[tracker_ind].NumberPole.Value)):
            sheet['A{0}'.format(row)] = row - 1
            sheet['C{0}'.format(row)] = pole_ind + 1
            sheet['D{0}'.format(row)] = poles[1 + pole_ind].Placement.Base.x * scale
            sheet['E{0}'.format(row)] = poles[1 + pole_ind].Placement.Base.y * scale
            sheet['F{0}'.format(row)] = poles[1 + pole_ind].Placement.Base.z * scale # buscar el z real (tocando suelo)
            style_range(sheet, 'A' + str(row) + ':F' + str(row),
                        border=Border(top=thin, left=thin, right=thin, bottom=thin),
                        alignment=Alignment(horizontal="center", vertical="center"))
            row += 1
        style_range(sheet, 'A' + str(group_from) + ':F' + str(row - 1),
                    border=Border(top=thin, left=thin, right=thin, bottom=thin),
                    alignment=Alignment(horizontal="center", vertical="center"))
        sheet.merge_cells('B' + str(group_from) + ':B' + str(row - 1))
        cell = sheet['B' + str(group_from)]
        cell.value = sel[tracker_ind].Label
        style_range(sheet, 'B' + str(group_from) + ':B' + str(row - 1),
                    border = Border(top = thin, left = thin, right = thin, bottom = thin),
                    alignment = Alignment(horizontal="center", vertical="center"))

        sheet['A{0}'.format(row)] = ""
        sheet.row_dimensions[row].height = 5
        row += 1

    mywb.save(filename)


    return

    import Spreadsheet
    i0 = '{0}'
    f3 = '{:.3f}'
    scale = 0.001

    sheet = FreeCAD.ActiveDocument.addObject("Spreadsheet::Sheet","BOQPoles")
    sheet.Label = "BOQ Poles"
    sheet.set('A1', 'Índice')
    sheet.set('B1', 'Tracker')
    sheet.set('C1', 'Hinca')
    sheet.set('D1', 'X')
    sheet.set('E1', 'Y')
    sheet.set('F1', 'Z')
    sheet.set('G1', 'Z')
    sheet.set('H1', 'Z')
    sheet.setStyle('A1:ZZ1', 'bold', 'add')
    sheet.setBackground('A1:H1', (0.5, 0.5, 0.5))

    row = 2
    for tracker_ind in range(1, len(sel)):
        poles = sel[tracker_ind].Shape.SubShapes
        group_from = row
        for pole_ind in range(0, int(sel[tracker_ind].NumberPole.Value)):
            sheet.set('A{0}'.format(row), str(row - 1))
            sheet.set('C{0}'.format(row), i0.format(pole_ind + 1))
            sheet.set('D{0}'.format(row), f3.format(poles[1 + pole_ind].Placement.Base.x * scale))
            sheet.set('E{0}'.format(row), f3.format(poles[1 + pole_ind].Placement.Base.y * scale))
            sheet.set('F{0}'.format(row), f3.format(poles[1 + pole_ind].Placement.Base.z * scale)) # buscar el z real (tocando suelo)
            row += 1

        sheet.set('A' + str(group_from) + ':F' + str(group_from))
        sheet.mergeCells('B' + str(group_from) + ':B' + str(row - 1))
        sheet.set('B{0}'.format(row - 1), sel[tracker_ind].Label)
    sheet.recompute()



def style_range(ws, cell_range, border=Border(), fill = None, font = None, alignment = None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def spreadsheetBOQPanelCollision(sel):
    import os
    path = os.path.dirname(FreeCAD.ActiveDocument.FileName)
    filename = os.path.join(path, "frames collision report.xlsx")

    i0 = '{0}'
    f3 = '{:.3f}'
    scale = 0.001

    mywb = openpyxl.Workbook()
    sheet = mywb.active
    sheet.title = 'Report'

    sheet['A1'] = 'Índice'
    sheet['B1'] = 'Nombre'
    sheet['C1'] = 'X'
    sheet['D1'] = 'Y'
    sheet['E1'] = 'Z'
    sheet['G1'] = 'Ángulo E-O'
    sheet['H1'] = 'Nº Hincas'

    for frame_ind in range(0, len(sel)):
        frame = sel[frame_ind]



        sheet['A{0}'.format(ind + 2)] = ind
        sheet['B{0}'.format(ind + 2)] = sel[ind].Label
        sheet['C{0}'.format(ind + 2)] = sel[ind].Placement.Base.x * scale
        sheet['D{0}'.format(ind + 2)] = sel[ind].Placement.Base.y * scale
        sheet['E{0}'.format(ind + 2)] = sel[ind].Placement.Base.z * scale
        sheet['F{0}'.format(ind + 2)] = sel[ind].Placement.Rotation.toEuler()[0]
        sheet['G{0}'.format(ind + 2)] = sel[ind].Placement.Rotation.toEuler()[1]
        sheet['H{0}'.format(ind + 2)] = sel[ind].NumberPole.Value
    mywb.save(filename)




class _CommandBOQFrames:

    def GetResources(self):
        return {'Pixmap': str(os.path.join(PVPlantResources.DirIcons, "boqm.svg")),
                'Accel': "R, M",
                'MenuText': QT_TRANSLATE_NOOP("Placement", "BOQ Mecánico"),
                'ToolTip': QT_TRANSLATE_NOOP("Placement", "Calcular el BOQ de la")}

    def Activated(self):
        sel = FreeCAD.ActiveDocument.findObjects(Name="Tracker")
        if len(sel) > 0:
            spreadsheetBOQFrames(sel)
            spreadsheetBOQPoles(sel)

    def IsActive(self):
        if FreeCAD.ActiveDocument:
            return True
        else:
            return False

if FreeCAD.GuiUp:
    FreeCADGui.addCommand('PVPlantBOQMechanical', _CommandBOQFrames())